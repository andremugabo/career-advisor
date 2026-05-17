import os
import re
import json
import openpyxl

def compile_mappings():
    dataset_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(dataset_dir, 'Education, Training, and Experience Categories.xlsx')
    sql_path = os.path.join(dataset_dir, '12_education_training_experience.sql')
    output_path = os.path.join(dataset_dir, 'education_experience_mappings.json')

    print("Parsing Excel Category Definitions...")
    # 1. Parse Excel definitions
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    # Scale mappings: (element_id, scale_id) -> category_id -> description
    category_definitions = {}
    
    for r in range(2, sheet.max_row + 1):
        elem_id = sheet.cell(row=r, column=1).value
        scale_id = sheet.cell(row=r, column=3).value
        category_id = sheet.cell(row=r, column=5).value
        description = sheet.cell(row=r, column=6).value
        
        if elem_id and scale_id and category_id is not None:
            key = f"{elem_id.strip()}_{scale_id.strip()}"
            if key not in category_definitions:
                category_definitions[key] = {}
            category_definitions[key][int(category_id)] = description.strip()

    print("Successfully parsed categories:", category_definitions.keys())

    # 2. Parse INSERT statements from SQL file
    # Example statement:
    # INSERT INTO education_training_experience (onetsoc_code, element_id, scale_id, category, data_value, ...) VALUES ('11-1011.00', '2.D.1', 'RL', 6, 21.61, ...);
    insert_pattern = re.compile(
        r"INSERT INTO education_training_experience\s*\([^)]+\)\s*VALUES\s*\('([^']+)',\s*'([^']+)',\s*'([^']+)',\s*([0-9.]+),\s*([0-9.]+)"
    )

    print("Scanning SQL records (this may take 10-15 seconds)...")
    # Store dynamic category values: onet_code -> key -> {category_id: data_value}
    career_metrics = {}

    if os.path.exists(sql_path):
        with open(sql_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                match = insert_pattern.search(line)
                if match:
                    code, elem_id, scale_id, cat_str, val_str = match.groups()
                    category_id = int(float(cat_str))
                    data_value = float(val_str)
                    
                    key = f"{elem_id}_{scale_id}"
                    # We only care about Education (2.D.1_RL), Experience (3.A.1_RW), and On-the-Job (3.A.3_OJ)
                    if key in ['2.D.1_RL', '3.A.1_RW', '3.A.3_OJ']:
                        if code not in career_metrics:
                            career_metrics[code] = {}
                        if key not in career_metrics[code]:
                            career_metrics[code][key] = {}
                        career_metrics[code][key][category_id] = data_value
    else:
        print(f"Error: SQL file not found at {sql_path}")
        return

    print(f"Parsed records for {len(career_metrics)} unique careers.")

    # 3. For each career, choose the category with the highest data_value percentage
    final_mappings = {}
    for code, metrics in career_metrics.items():
        career_entry = {}
        
        # 2.D.1 Required level of education
        edu_metrics = metrics.get('2.D.1_RL', {})
        if edu_metrics:
            best_cat = max(edu_metrics, key=edu_metrics.get)
            career_entry['required_education'] = category_definitions.get('2.D.1_RL', {}).get(best_cat, "Unknown")
        else:
            career_entry['required_education'] = "Not Specified"
            
        # 3.A.1 Related work experience
        exp_metrics = metrics.get('3.A.1_RW', {})
        if exp_metrics:
            best_cat = max(exp_metrics, key=exp_metrics.get)
            career_entry['work_experience'] = category_definitions.get('3.A.1_RW', {}).get(best_cat, "Unknown")
        else:
            career_entry['work_experience'] = "Not Specified"

        # 3.A.3 On-the-job training
        oj_metrics = metrics.get('3.A.3_OJ', {})
        if oj_metrics:
            best_cat = max(oj_metrics, key=oj_metrics.get)
            career_entry['on_the_job_training'] = category_definitions.get('3.A.3_OJ', {}).get(best_cat, "Unknown")
        else:
            career_entry['on_the_job_training'] = "Not Specified"
            
        final_mappings[code] = career_entry

    print(f"Saving compiled dataset to {output_path}...")
    with open(output_path, 'w', encoding='utf-8') as out_f:
        json.dump(final_mappings, out_f, indent=2)
    print("Compilation Complete!")

if __name__ == '__main__':
    compile_mappings()
